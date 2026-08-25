from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]

INPUT = ROOT / "OmniRoute_Provider_Intelligence_Master_Updated_2026-08-20_Batch86.xlsx"
OUTPUT = ROOT / "open-sse/config/providerIntelligence.generated.ts"


from datetime import date, datetime, timedelta


NUMERIC_FIELDS = {
    "Model Count",
    "Free Amount",
    "Overall Score",
    "Amount",
    "RPM",
    "RPH",
    "RPD",
    "TPM",
    "Concurrency",
}


def excel_serial_to_iso(value):
    # Excel's 1900 date system. openpyxl may expose dates as numeric serials
    # when workbook formatting is inconsistent.
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 30000 <= float(value) <= 60000:
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=float(value))
            return dt.date().isoformat()
    return None


def clean_field(name, value):
    if value is None:
        return None

    field = str(name).strip()

    if field == "Last Verified":
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        converted = excel_serial_to_iso(value)
        if converted is not None:
            return converted
        return str(value)

    if field in NUMERIC_FIELDS:
        if isinstance(value, bool):
            return None

        if isinstance(value, (int, float)):
            return int(value) if float(value).is_integer() else float(value)

        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            if not stripped:
                return None
            try:
                number = float(stripped)
                return int(number) if number.is_integer() else number
            except ValueError:
                return None

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


def row_to_dict(headers, values):
    return {
        str(h).strip(): clean_field(h, v)
        for h, v in zip(headers, values)
        if h is not None
    }


def main():
    if not INPUT.exists():
        raise SystemExit(f"Input workbook not found: {INPUT}")

    wb = load_workbook(INPUT, data_only=True)

    provider_ws = wb["Provider Master"]
    capacity_ws = wb["Free Capacity"]

    provider_headers = [c.value for c in provider_ws[1]]
    capacity_headers = [c.value for c in capacity_ws[1]]

    provider_rows = [
        row_to_dict(provider_headers, row)
        for row in provider_ws.iter_rows(min_row=2, values_only=True)
    ]

    capacity_rows = [
        row_to_dict(capacity_headers, row)
        for row in capacity_ws.iter_rows(min_row=2, values_only=True)
    ]

    # Preserve duplicate provider IDs such as k3 by indexing on
    # (Provider ID, Alias), not Provider ID alone.
    capacity_index = {}
    for row in capacity_rows:
        key = (
            row.get("Provider"),
            None,  # Free Capacity sheet currently identifies by Provider.
        )
        capacity_index[key] = row

    entries = []

    for row in provider_rows:
        provider_id = row.get("Provider ID")
        alias = row.get("Alias")

        if not provider_id:
            continue

        capacity = capacity_index.get((provider_id, None), {})

        entries.append(
            {
                "providerId": provider_id,
                "alias": alias,
                "format": row.get("Format"),
                "executor": row.get("Executor"),
                "authType": row.get("Auth Type"),
                "oauth": row.get("OAuth"),
                "anonymous": row.get("Anonymous"),
                "passthrough": row.get("Passthrough"),
                "modelCount": row.get("Model Count"),
                "baseUrl": row.get("Base URL"),
                "registryFreeSignal": row.get("Registry Free Signal"),
                "freeType": row.get("Free Type"),
                "freeAmount": row.get("Free Amount"),
                "freeUnit": row.get("Free Unit"),
                "resetPeriod": row.get("Reset Period"),
                "accountRequired": row.get("Account Required"),
                "paymentRequired": row.get("Payment Required"),
                "tosRisk": row.get("ToS Risk"),
                "tier": row.get("Tier"),
                "decision": row.get("Decision"),
                "overallScore": row.get("Overall Score"),
                "confidence": row.get("Confidence"),
                "lastVerified": row.get("Last Verified"),
                "notes": row.get("Notes"),
                "freeCapacity": {
                    "amount": capacity.get("Amount"),
                    "unit": capacity.get("Unit"),
                    "scope": capacity.get("Scope"),
                    "resetPeriod": capacity.get("Reset Period"),
                    "resetDetails": capacity.get("Reset Details"),
                    "rpm": capacity.get("RPM"),
                    "rph": capacity.get("RPH"),
                    "rpd": capacity.get("RPD"),
                    "tpm": capacity.get("TPM"),
                    "concurrency": capacity.get("Concurrency"),
                    "timeRestrictions": capacity.get("Time Restrictions"),
                    "accountRequired": capacity.get("Account Required"),
                    "paymentRequired": capacity.get("Payment Required"),
                    "modelRestrictions": capacity.get("Model Restrictions"),
                    "registryEvidence": capacity.get("Registry Evidence"),
                    "confidence": capacity.get("Confidence"),
                },
            }
        )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    header = """/**
 * AUTO-GENERATED from OmniRoute Provider Intelligence Master.
 * Do not edit by hand.
 *
 * Source workbook:
 * OmniRoute_Provider_Intelligence_Master_Updated_2026-08-20_Batch86.xlsx
 */

export type ProviderIntelligenceTier =
  | "Primary"
  | "Secondary"
  | "Opportunistic"
  | "Promotional"
  | "Experimental"
  | "Exclude"
  | "Specialized"
  | "Unclassified";

export interface ProviderFreeCapacity {
  amount: number | null;
  unit: string | null;
  scope: string | null;
  resetPeriod: string | null;
  resetDetails: string | null;
  rpm: number | null;
  rph: number | null;
  rpd: number | null;
  tpm: number | null;
  concurrency: number | null;
  timeRestrictions: string | null;
  accountRequired: string | null;
  paymentRequired: string | null;
  modelRestrictions: string | null;
  registryEvidence: string | null;
  confidence: string | null;
}

export interface ProviderIntelligence {
  providerId: string;
  alias: string | null;
  format: string | null;
  executor: string | null;
  authType: string | null;
  oauth: string | null;
  anonymous: string | null;
  passthrough: string | null;
  modelCount: number | null;
  baseUrl: string | null;
  registryFreeSignal: string | null;
  freeType: string | null;
  freeAmount: number | null;
  freeUnit: string | null;
  resetPeriod: string | null;
  accountRequired: string | null;
  paymentRequired: string | null;
  tosRisk: string | null;
  tier: ProviderIntelligenceTier;
  decision: string | null;
  overallScore: number | null;
  confidence: string | null;
  lastVerified: string | null;
  notes: string | null;
  freeCapacity: ProviderFreeCapacity;
}

export const PROVIDER_INTELLIGENCE_CURATED_AT = "2026-08-20";

export const PROVIDER_INTELLIGENCE_SOURCE =
  "OmniRoute_Provider_Intelligence_Master_Updated_2026-08-20_Batch86.xlsx";

export const PROVIDER_INTELLIGENCE: readonly ProviderIntelligence[] =
"""

    body = json.dumps(entries, ensure_ascii=False, indent=2)

    footer = """;

const BY_KEY = new Map<string, ProviderIntelligence>();

for (const entry of PROVIDER_INTELLIGENCE) {
  const key = entry.alias
    ? `${entry.providerId}:${entry.alias}`
    : entry.providerId;

  BY_KEY.set(key, entry);

  if (!BY_KEY.has(entry.providerId)) {
    BY_KEY.set(entry.providerId, entry);
  }
}

export function getProviderIntelligence(
  providerId: string,
  alias?: string | null,
): ProviderIntelligence | null {
  if (alias) {
    return BY_KEY.get(`${providerId}:${alias}`) ?? null;
  }

  return BY_KEY.get(providerId) ?? null;
}

export function isFreeCandidate(
  providerId: string,
  alias?: string | null,
): boolean {
  const entry = getProviderIntelligence(providerId, alias);
  if (!entry) return false;

  return (
    entry.tier === "Primary" ||
    entry.tier === "Secondary" ||
    entry.tier === "Opportunistic"
  );
}

export function isExcludedProvider(
  providerId: string,
  alias?: string | null,
): boolean {
  return getProviderIntelligence(providerId, alias)?.tier === "Exclude";
}
"""

    OUTPUT.write_text(header + body + footer, encoding="utf-8")

    print(f"Generated: {OUTPUT}")
    print(f"Provider Master rows: {len(provider_rows)}")
    print(f"Generated entries: {len(entries)}")


if __name__ == "__main__":
    main()
